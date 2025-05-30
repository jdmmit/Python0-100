import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# Crear un archivo Excel en memoria
output = io.BytesIO()

# Crear un libro de trabajo
wb = Workbook()

# 1. Lista de chequeo de requisitos
ws1 = wb.active
ws1.title = "Lista de Chequeo Requisitos"

# Definir estilos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Encabezados
headers = ["ID", "Descripción del Requisito", "Criterios de Aceptación", "Prioridad", "Estado", "Responsable", "Observaciones"]
for col_num, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Ejemplo de datos
example_data = [
    ["REQ-001", "El sistema debe permitir el registro de usuarios", "El usuario puede registrarse con email y contraseña", "Alta", "Aprobado", "Juan Pérez", ""],
    ["REQ-002", "El sistema debe permitir la recuperación de contraseña", "El usuario recibe un email con instrucciones", "Media", "Pendiente", "María López", ""],
    ["REQ-003", "El sistema debe mostrar un dashboard con estadísticas", "Se muestran gráficos de uso y actividad", "Baja", "Implementado", "Carlos Gómez", ""]
]

for row_num, row_data in enumerate(example_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_num, column=col_num, value=cell_value)
        cell.border = border
        if col_num == 4:  # Columna de Prioridad
            if cell_value == "Alta":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif cell_value == "Media":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif cell_value == "Baja":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Ajustar ancho de columnas
for col in range(1, len(headers) + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 20

# 2. Plantilla de casos de prueba
ws2 = wb.create_sheet(title="Casos de Prueba")

# Encabezados
headers = ["ID", "Descripción", "Precondiciones", "Pasos", "Resultado Esperado", "Estado", "Ejecutado por", "Fecha"]
for col_num, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Ejemplo de datos
example_data = [
    ["CP-001", "Verificar registro de usuario", "Aplicación iniciada", "1. Ir a registro\n2. Ingresar datos\n3. Hacer clic en 'Registrar'", "Usuario registrado exitosamente", "Aprobado", "Ana Martínez", "2023-05-15"],
    ["CP-002", "Verificar login de usuario", "Usuario registrado", "1. Ir a login\n2. Ingresar credenciales\n3. Hacer clic en 'Ingresar'", "Usuario ingresa al sistema", "Fallido", "Pedro Sánchez", "2023-05-16"],
    ["CP-003", "Verificar recuperación de contraseña", "Usuario registrado", "1. Ir a 'Olvidé mi contraseña'\n2. Ingresar email\n3. Hacer clic en 'Recuperar'", "Email enviado con instrucciones", "Pendiente", "", ""]
]

for row_num, row_data in enumerate(example_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_num, column=col_num, value=cell_value)
        cell.border = border
        if col_num == 4:  # Columna de Pasos
            cell.alignment = Alignment(wrap_text=True)
        if col_num == 6:  # Columna de Estado
            if cell_value == "Aprobado":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif cell_value == "Fallido":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif cell_value == "Pendiente":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Ajustar ancho de columnas
for col in range(1, len(headers) + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 20
ws2.column_dimensions[get_column_letter(4)].width = 30
ws2.column_dimensions[get_column_letter(5)].width = 30
ws2.row_dimensions[2].height = 60
ws2.row_dimensions[3].height = 60
ws2.row_dimensions[4].height = 60

# 3. Formulario de revisión de código
ws3 = wb.create_sheet(title="Revisión de Código")

# Encabezados
headers = ["ID", "Archivo", "Lenguaje", "Criterio", "Cumple (Sí/No)", "Comentarios", "Revisor", "Fecha"]
for col_num, header in enumerate(headers, 1):
    cell = ws3.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Ejemplo de datos
example_data = [
    ["RC-001", "login.php", "PHP", "Legibilidad", "Sí", "Código bien comentado", "Luis Ramírez", "2023-05-20"],
    ["RC-002", "login.php", "PHP", "Seguridad", "No", "No valida entrada de usuario", "Luis Ramírez", "2023-05-20"],
    ["RC-003", "dashboard.js", "JavaScript", "Eficiencia", "Sí", "Optimizado para carga rápida", "Carmen Díaz", "2023-05-21"]
]

for row_num, row_data in enumerate(example_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_num, column=col_num, value=cell_value)
        cell.border = border
        if col_num == 5:  # Columna de Cumple
            if cell_value == "Sí":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif cell_value == "No":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Ajustar ancho de columnas
for col in range(1, len(headers) + 1):
    ws3.column_dimensions[get_column_letter(col)].width = 15
ws3.column_dimensions[get_column_letter(6)].width = 30

# 4. Encuesta de satisfacción del cliente (formato)
ws4 = wb.create_sheet(title="Encuesta Satisfacción")

# Título
ws4.merge_cells('A1:G1')
cell = ws4.cell(row=1, column=1, value="ENCUESTA DE SATISFACCIÓN DEL CLIENTE")
cell.font = Font(bold=True, size=14)
cell.alignment = Alignment(horizontal='center', vertical='center')

# Información del cliente
ws4.merge_cells('A3:B3')
ws4.cell(row=3, column=1, value="Nombre del Cliente:").font = Font(bold=True)
ws4.merge_cells('C3:G3')

ws4.merge_cells('A4:B4')
ws4.cell(row=4, column=1, value="Proyecto:").font = Font(bold=True)
ws4.merge_cells('C4:G4')

ws4.merge_cells('A5:B5')
ws4.cell(row=5, column=1, value="Fecha:").font = Font(bold=True)
ws4.merge_cells('C5:G5')

# Instrucciones
ws4.merge_cells('A7:G7')
cell = ws4.cell(row=7, column=1, value="Por favor, califique los siguientes aspectos del software (1: Muy insatisfecho, 5: Muy satisfecho)")
cell.font = Font(italic=True)
cell.alignment = Alignment(horizontal='center', vertical='center')

# Encabezados de la encuesta
headers = ["Aspecto", "1", "2", "3", "4", "5", "Comentarios"]
for col_num, header in enumerate(headers, 1):
    cell = ws4.cell(row=9, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Aspectos a evaluar
aspects = [
    "Funcionalidad (cumple con los requisitos)",
    "Usabilidad (facilidad de uso)",
    "Rendimiento (velocidad, eficiencia)",
    "Confiabilidad (estabilidad, pocos errores)",
    "Soporte técnico",
    "Documentación",
    "Satisfacción general"
]

for row_num, aspect in enumerate(aspects, 10):
    cell = ws4.cell(row=row_num, column=1, value=aspect)
    cell.border = border
    for col_num in range(2, 8):
        cell = ws4.cell(row=row_num, column=col_num)
        cell.border = border

# Preguntas abiertas
row_num = len(aspects) + 11
ws4.merge_cells(f'A{row_num}:G{row_num}')
cell = ws4.cell(row=row_num, column=1, value="¿Qué aspectos del software le gustaron más?")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal='left', vertical='center')

row_num += 1
ws4.merge_cells(f'A{row_num}:G{row_num+2}')
cell = ws4.cell(row=row_num, column=1)
cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

row_num += 4
ws4.merge_cells(f'A{row_num}:G{row_num}')
cell = ws4.cell(row=row_num, column=1, value="¿Qué aspectos del software cree que podrían mejorarse?")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal='left', vertical='center')

row_num += 1
ws4.merge_cells(f'A{row_num}:G{row_num+2}')
cell = ws4.cell(row=row_num, column=1)
cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Ajustar ancho de columnas
ws4.column_dimensions['A'].width = 40
for col in range(2, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 5
ws4.column_dimensions['G'].width = 30

# Guardar el archivo
wb.save(output)
output.seek(0)

# Crear un archivo Excel para descargar
with open('Instrumentos_Calidad_Software.xlsx', 'wb') as f:
    f.write(output.getvalue())

print("Archivo Excel 'Instrumentos_Calidad_Software.xlsx' creado exitosamente con los siguientes instrumentos:")
print("1. Lista de Chequeo de Requisitos")
print("2. Plantilla de Casos de Prueba")
print("3. Formulario de Revisión de Código")
print("4. Encuesta de Satisfacción del Cliente")